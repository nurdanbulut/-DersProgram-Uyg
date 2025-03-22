import string
import random
from pathlib import Path
from uuid import uuid4

from flask import request
from openpyxl.reader.excel import load_workbook

from peewee import fn, JOIN

from openpyxl import load_workbook

from .models import Department, User, SyllabusOptions, Lesson, Syllabus, LessonType, Excel

def generate_password(length:int=12):
    """Belirtilen uzunlukta rastgele bir şifre üretir."""
    characters = string.ascii_letters + string.digits + string.punctuation
    return ''.join(random.choices(characters, k=length))

def get_student_count_of_department(deparment:Department)->int:
    """
    Departmana ait öğrenci sayısını belirtir
    :param deparment: Departman objesi
    :return: Öğrenci sayısı
    """

    return User.select().where(User.department == deparment, User.type == 0).count()

def parse_options(request:request) -> list:
    """
    SyllabusOptions Eklerken aynı tipten birden fazla veri alındığı için verileri parse etmek için kullanılıyor.
    Request parametresine gelen requesti veriyorsunuz ve ardından size parse edilmiş bir liste döndürüyor.
    Liste şu şekilde olması gerekiyor [LessonType_id, Hours]
    İlk parametre ders tipinin id numarası, İkinci parametre ise saat
    """
    result = []

    index = 0
    while lesson := request.form.get(f'lesson-{index}', None, type=int):
        result.append((
            lesson,
            request.form.get(f'hour-{index}', 1, type=int),
            request.form.get(f'shareable-{index}', False, type=bool)
        ))

        index += 1

    return result

def format_options(department, grade):
    result = []

    for option in SyllabusOptions.select(SyllabusOptions.lesson_type, SyllabusOptions.hpw, SyllabusOptions.shareable).where(SyllabusOptions.department==department, SyllabusOptions.grade==grade):
        result.append([option.lesson_type.id, option.hpw, option.shareable])

    return result


def get_shared_lessons() -> list[SyllabusOptions]:
    """
    Ortak dersleri çeken bir fonksiyon
    :return:
    """
    SyllabusOptionsAlias = SyllabusOptions.alias()

    query = (SyllabusOptions
             .select(SyllabusOptions, LessonType, Department)  # Tüm sütunları seç
             .join(SyllabusOptionsAlias, JOIN.INNER, on=(SyllabusOptions.lesson_type == SyllabusOptionsAlias.lesson_type))
             .join(LessonType, on=(SyllabusOptions.lesson_type == LessonType.id))  # LessonType ile join
             .join(Department, on=(SyllabusOptions.department == Department.id))  # Department ile join
             .where(SyllabusOptions.department != SyllabusOptionsAlias.department, SyllabusOptions.shareable==True)
             .group_by(SyllabusOptions.lesson_type, SyllabusOptions.department)
             .order_by(SyllabusOptions.hpw))
    return query

def place_lesson(syllabus_options:SyllabusOptions, syllabus:Syllabus, shared=False):
    """
    Verilen dersi müsaitlik durumuna göre kontrol eder ve yerleştirir.
    """
    if syllabus_options.lesson_type.online:
        hours = range(17, 19)
    else:
        hours = range(9, 17)

    day = 1
    hour_index = 0
    while True:
        # Öğretmen müsait mi kontrol et
        if lesson_check := (Lesson
                .select()
                .where(Lesson.teacher == syllabus_options.lesson_type.teacher,
                       Lesson.start_hour == hours[hour_index],
                       Lesson.day == day,
                       Lesson.syllabus == syllabus)
                .get_or_none()): # Öğretmen müsait değilse sonraki derse geç

            hour_index += 1
            if hour_index >= len(hours):  # Sonraki ders yoksa sonraki güne geç
                hour_index = 0
                day += 1
                if day == 6: # Sonraki gün yoksa hata ver
                    return (False, f"{lesson_check.teacher.name} isimli öğretmenin ders ayarlarını düzenleyin")

            continue # Eğer müsait değilse döngüyü devam ettir
        # Bölümün o tarihte dersi var mı kontrol et
        if lesson_check := (Lesson
                .select()
                .join(SyllabusOptions)
                .where(Lesson.start_hour == hours[hour_index],
                       Lesson.day == day,
                       Lesson.grade == syllabus_options.grade,
                       Lesson.options.department == syllabus_options.department,
                       Lesson.syllabus == syllabus)
                .get_or_none()): # Eğer dersi varsa sonraki derse geç

            hour_index += 1
            if hour_index >= len(hours):# Sonraki ders yoksa sonraki güne geç
                hour_index = 0
                day += 1
                # Burada ekstra bir kontrol yok çünkü maksimum haftalık saati geçemezler zaten
            continue # Eğer müsait değilse döngüyü devam ettir

        # Dersi ekle
        lesson = Lesson(
            day=day,
            start_hour=hours[hour_index],
            type=syllabus_options.lesson_type,
            teacher=syllabus_options.lesson_type.teacher,
            grade=syllabus_options.grade,
            options=syllabus_options,
            syllabus=syllabus,
            shared=shared
        )
        lesson.save()

        return (True, lesson)

def get_lessons(syllabus, department, grade):
    table = [
        [None for _ in range(11)] for _ in range(5)
    ]

    for lesson in (Lesson
            .select()
            .join(SyllabusOptions)
            .where(Lesson.syllabus == syllabus,
                   Lesson.options.department == department,
                   Lesson.grade == grade)):
        table[lesson.day - 1][lesson.start_hour - 9] = lesson
    return table

def get_lessons_for_teacher(syllabus, teacher):
    table = [
        [[] for _ in range(11)] for _ in range(5)
    ]

    for lesson in (Lesson
            .select()
            .join(SyllabusOptions)
            .join(LessonType)
            .where(Lesson.syllabus == syllabus,
                   Lesson.type.teacher == teacher)):
        if lesson:
            table[lesson.day - 1][lesson.start_hour - 9].append(lesson)
    return table

days = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma"]

get_day = lambda day: days[day - 1]


excel_locations = (
    # Sütun, sıra
    ((3, 4), (3, 5), (3, 6), (3, 7), (3, 8), (3, 9), (3, 10), (3, 11), (3, 12), (3, 13)),
    ((3, 14), (3, 15), (3, 16), (3, 17), (3, 18), (3, 19), (3, 20), (3, 21), (3, 22), (3, 23)),
    ((3, 24), (3, 25), (3, 26), (3, 27), (3, 28), (3, 29), (3, 30), (3, 31), (3, 32), (3, 33)),
    ((3, 34), (3, 35), (3, 36), (3, 37), (3, 38), (3, 39), (3, 40), (3, 41), (3, 42), (3, 43)),
    ((3, 44), (3, 45), (3, 46), (3, 47), (3, 48), (3, 49), (3, 50), (3, 51), (3, 52), (3, 53))
)

def create_excel(excel_folder:Path, folder_name, syllabus:Syllabus, department:Department):
    file = load_workbook((excel_folder / 'template.xlsx').resolve())
    ws = file.active

    ws.cell(row=1, column=3, value=f"{department.name} : {department.short_name}")

    for lesson in Lesson.select().join(SyllabusOptions).where(Lesson.syllabus == syllabus, Lesson.options.department == department):
        col, row = excel_locations[lesson.day - 1][lesson.start_hour - 9]

        col += (lesson.options.grade - 1) # sınıfın bir eksiğini ekliyorum ki indexe uyum sağlasın

        ws.cell(
            row=row,
            column=col,
            value=str(lesson.type),
        )

    filename = str(uuid4()) + '.xlsx'
    file.save((excel_folder / 'output' / folder_name / filename).resolve())

    Excel(
        department=department,
        filename=filename,
        syllabus=syllabus
    ).save()